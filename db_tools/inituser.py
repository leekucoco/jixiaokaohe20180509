import sys
import os
from openpyxl import load_workbook,Workbook

from openpyxl.worksheet.table import Table, TableStyleInfo


from datetime import datetime,date
pwd = os.path.dirname(os.path.realpath(__file__))
sys.path.append(pwd)
sys.path.append("..")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "Dqrcbankjxkh.settings")
import django
django.setup()
import xlwt,openpyxl
from django.contrib.auth import get_user_model
from users.models import *
User = get_user_model()
from certificates.models import *
from depart.models import *
from rank13.models import *
from coefficient.models import *

def initsuperuser():
    u = User.objects.get(id = 1)
    u.set_password("123456")
    u.save()

def parselevel(choices,info):
    res = 1
    for t in choices:
        #print(t,info,t[1],info in t[1],t[0])
        if info in t[1]:
            res = t[0]
        else:
            continue
    return res


def parseexcelline(filename):
    wb = load_workbook(filename)
    sheet = wb.active
    sheetdimensions = sheet.dimensions
    for row in sheet[sheetdimensions]:
        l2 = []
        for cell in row:
            t = cell.value
            l2.append(t)
        else:
            t = ""
            l2.append(t)
        yield l2


def writelinetoexcel(ls, filename):
    wb = Workbook()
    sheet = wb.active
    for line in ls:
        sheet.append([line])
    wb.save(filename)


def quchonglist(index, line):
    temp = []
    for elem in line:
        info = elem[index]
        if info not in temp:
            temp.append(info)
        else:
            pass
    return temp


def initusers():
    line = parseexcelline("table20180725.xlsx")
    count = 0
    dic={}
    for elem in line:

        if elem[22] != None and elem[11] != None:
            dic["username"] = elem[22]
            datinfo = elem[24]
            dat = datinfo.split(".")
            if dat != "":
                if len(dat) == 2:
                    stdate = date(int(dat[0]), int(dat[1]), 1)
                elif len(dat) == 1:
                    stdate = date(int(dat[0]), 1, 1)
                else:
                    stdate = date.today()
            else:
                stdate = ""

            dic["joinedyears"] = stdate
            dic["idcardnumber"] = elem[0]

            dic["name"] = elem[1]
            dic["mobile"] = ""
            dic["clerkrank"] = parselevel(UserProfile.CLERKRANK_CHONCES,elem[6])
            dic["cmanagerrank"] = parselevel(UserProfile.CMANAGERRANK_CHOICES,elem[7])
            dic["cmanagerlevel"] = parselevel(UserProfile.CMANAGERLEVEL_CHOICES,elem[8])
            dic["education"] = parselevel(UserProfile.EDUCATION_CHOICES,elem[11])

            dic["title"] = parselevel(UserProfile.TITLE_CHOICES,elem[12])
            if elem[13] != "nan":
                dic["primccbp"] = int(elem[13])
            else:
                dic["primccbp"] = 0
            if elem[14] != "nan":
                dic["intermediateccbp"] = int(elem[14])
            else:
                dic["intermediateccbp"] = 0

            dic["internel_trainer"] = parselevel(UserProfile.INTERNEL_TRAINER_CHOICES,elem[15])
            # print(dic)
            try:
                User.objects.create(**dic)

                count = count +1

            except Exception as e:
                print(dic,e)
        else:
            pass


        #print(dic)
    print("成功生成用户%d"%count)



def updateuserbaseinfo():
    line = parseexcelline("table20180725.xlsx")
    count = 0
    dic={}
    for elem in line:
        if elem[22] != None and elem[11] != None:
            datinfo = elem[10]
            dat = datinfo.split(".")
            if dat != "":
                if len(dat) == 2:
                    stdate = date(int(dat[0]), int(dat[1]), 1)
                elif len(dat) == 1:
                    stdate = date(int(dat[0]), 1, 1)
                else:
                    stdate = date.today()
            else:
                stdate = ""

            datworkingyears = elem[24]
            datw = datworkingyears.split(".")
            if datw != "":
                if len(datw) == 2:
                    stdatew = date(int(datw[0]), int(datw[1]), 1)
                elif len(datw) == 1:
                    stdatew = date(int(datw[0]), 1, 1)
                else:
                    stdatew = date.today()
            else:
                stdatew = ""

            dic["workingyears"] = stdatew
            dic["joinedyears"] = stdate
            dic["idcardnumber"] = elem[0]
            if int(elem[0][16]) %2 == 1:
                dic["gender"] = "male"
            else:
                dic["gender"] = "female"
            dic["mobile"] = ""
            dic["clerkrank"] = parselevel(UserProfile.CLERKRANK_CHONCES,elem[6])
            dic["cmanagerrank"] = parselevel(UserProfile.CMANAGERRANK_CHOICES,elem[7])
            dic["cmanagerlevel"] = parselevel(UserProfile.CMANAGERLEVEL_CHOICES,elem[8])
            dic["education"] = parselevel(UserProfile.EDUCATION_CHOICES,elem[11])

            dic["title"] = parselevel(UserProfile.TITLE_CHOICES,elem[12])
            if elem[13] != "nan":
                dic["primccbp"] = int(elem[13])
            else:
                dic["primccbp"] = 0
            if elem[14] != "nan":
                dic["intermediateccbp"] = int(elem[14])
            else:
                dic["intermediateccbp"] = 0

            dic["internel_trainer"] = parselevel(UserProfile.INTERNEL_TRAINER_CHOICES,elem[15])
            dic["addbasesalary"] = elem[9]
            try:
                u = User.objects.get(idcardnumber=dic["idcardnumber"])
                #print(u)
                if u.education != dic["education"] or u.title != dic["title"]\
                        or u.internel_trainer != dic["internel_trainer"] or u.clerkrank != dic["clerkrank"]\
                        or u.cmanagerrank != dic["cmanagerrank"] or u.cmanagerlevel != dic["cmanagerlevel"] \
                        or u.workingyears != dic["workingyears"] or u.joinedyears != dic["joinedyears"]\
                        or u.gender != dic["gender"] or u.addbasesalary != dic["addbasesalary"]:
                    u.education = dic["education"]
                    #print(u.education)
                    u.title = dic["title"]
                    u.internel_trainer = dic["internel_trainer"]
                    u.clerkrank = dic["clerkrank"]
                    u.cmanagerrank = dic["cmanagerrank"]
                    u.cmanagerlevel = dic["cmanagerlevel"]
                    u.workingyears = dic["workingyears"]
                    u.joinedyears = dic["joinedyears"]
                    u.gender = dic["gender"]
                    u.addbasesalary = dic["addbasesalary"]
                    count = count + 1
                    # print(count)
                    u.save()
                else:
                    pass
                pass
            except Exception:
                pass
    print("成功更新用基础信息%d"%count)

def inituserspassword():
    users = User.objects.all()
    for u in users:
        u.set_password("123456")
        u.save()
    print("完成密码重置任务")


def initcertificates():
    l = parseexcelline("certificates.xlsx")
    dic={}
    count = 0
    for it in l:
        dic["name"] = it[0]
        dic["score"] = it[1]
        dic["desc"] = it[0]
        # print(it)
        try:
            Cerficates.objects.create(**dic)
            count = count + 1
        except Exception:
            pass
    print("成功生产证书%d"%count)



def initusercertificates():
    line = parseexcelline("table20180725.xlsx")
    count = 0
    dic={}
    for elem in line:
        if elem[22] != None and elem[11] != None:
            username = elem[22]
            ce = [elem[16],elem[17],elem[18],elem[19],elem[20]]
            count = 0
            u = User.objects.get(username=username)
            for c in ce:
                cer = Cerficates.objects.filter(name=c)
                if cer:
                    dic["user"] = u
                    dic["certificate"] = cer[0]
                    try:
                        IndexUserCertificate.objects.create(**dic)
                        count = count + 1
                    except Exception:
                        pass
                else:
                    pass
               #print(dic)
    print("成功生成用户-证书%d"%count)




def initdeparts():
    wb = load_workbook("departanposts.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    l1 = []
    # f = open("log4.txt","a+")
    agent = Agent.objects.get(id=1)
    sheetdimensions = sheet.dimensions
    for row in sheet[sheetdimensions]:
        l2 =[]
        for cell in row:
            if cell.value != None:
                t = cell.value
                l2.append(t)
            else:
                t = ""
                l2.append(t)
        for d in l2:
            dic = {}
            dic["name"] = d
            dic["desc"] = d
            dic["agent"] = agent
            dic["dept_type"] = 1
            dic["basesalary"] = 1900
            DepartDetail.objects.create(**dic)


def initposts():
    wb = load_workbook("departanposts.xlsx")
    sheet = wb.get_sheet_by_name("Sheet2")
    l1 = []
    # f = open("log4.txt","a+")
    #agent = Agent.objects.get(id=1)
    sheetdimensions = sheet.dimensions
    for row in sheet[sheetdimensions]:
        l2 =[]
        for cell in row:
            if cell.value != None:
                t = cell.value
                l2.append(t)
            else:
                t = ""
                l2.append(t)
        for d in l2:
            dic = {}
            dic["name"] = d
            #print(dic)
            try:
                Post.objects.create(**dic)
            except Exception:
                pass



def initrank13demands():

    wb = load_workbook("rank13demandcoefficients.xlsx")
    sheet = wb.get_sheet_by_name("Sheet3")
    l1 = []
    for row in sheet.iter_rows('A1:K17'):
        l2 =[]
        for cell in row:
            if cell.value != None:
                t = cell.value
                l2.append(t)
            else:
                t = ""
                l2.append(t)
        #print(l2)

        posts = l2[1].split(" ")
        #print(len(posts))
        try:
            for post in posts:
                dic = {}
                dic["agent"] = Agent.objects.get(id = 1)
                dic["post"] = Post.objects.get(name=post)
                dic["rank"] = l2[0]
                dic["demandyears"] = l2[2]
                dic["educationdemands"] =parselevel(Rank13Demands.EDUCATION_CHOICES,l2[3])
                dic["primccbpdemands"] = l2[4]
                dic["titledemands"] = parselevel(Rank13Demands.TITLE_CHOICES,l2[5])
                #print(l2)
                #print(dic)
                try:
                    Rank13Demands.objects.create(**dic)
                except Exception:
                    pass
        except Exception:
            pass



def initrank13coe():

    wb = load_workbook("rank13demandcoefficients.xlsx")
    sheet = wb.get_sheet_by_name("Sheet3")
    l1 = []
    for row in sheet.iter_rows('A1:K17'):
        l2 =[]
        for cell in row:
            if cell.value != None:
                t = cell.value
                l2.append(t)
            else:
                t = ""
                l2.append(t)
        #print(l2)

        posts = l2[1].split(" ")
        #print(len(posts))
        try:
            for post in posts:
                dic = {}
                dic["agent"] = Agent.objects.get(id = 1)
                dic["post"] = Post.objects.get(name=post)
                dic["rank"] = l2[0]
                dic["demandyears"] = l2[2]
                dic["educationdemands"] =parselevel(Rank13Demands.EDUCATION_CHOICES,l2[3])
                dic["primccbpdemands"] = l2[4]
                dic["titledemands"] = parselevel(Rank13Demands.TITLE_CHOICES,l2[5])
                #print(l2)
                #print(dic)
                # try:
                #     Rank13Demands.objects.create(**dic)
                # except Exception:
                #     pass
                del dic["demandyears"]
                del dic["educationdemands"]
                del dic["primccbpdemands"]
                del dic["titledemands"]

                for i in range(1,6):
                    dic["level"] = i
                    dic["coefficent"] = l2[i+5]
                    #print(dic)
                    try:
                        Rank13Coefficent.objects.create(**dic)
                    except Exception:
                        pass

        except Exception:
            pass


def initcoe():
    line = parseexcelline("table20180725.xlsx")
    count = 0
    dic = {}
    for elem in line:
        if elem[22] != None and elem[11] != None:
            username = elem[22]
            postname = elem[4]
            rank = elem[5]
            rank13demands = Rank13Demands.objects.filter(post__name=postname,rank=rank)
            if rank13demands:
                user = User.objects.get(username=username)
                dic["user"] = user
                dic["rank13demands"] = rank13demands[0]
                try:
                    CoefficientDetail.objects.create(**dic)
                    count = count + 1
                except Exception:
                    continue
            else:
                print(elem)
    print("成功生产员工-系数%d"%count)


def inituserdepart():
    line = parseexcelline("table20180725.xlsx")
    count = 0
    dic={}
    for elem in line:
        if elem[22] != None and elem[11] != None:
            username = elem[22]
            departs = DepartDetail.objects.filter(name=elem[2])
            if departs:
                user = User.objects.get(username=username)
                dic["user"] = user
                dic["depart"] = departs[0]
                try:
                    IndexUserDepart.objects.create(**dic)
                except Exception:
                    pass
                count = count +1
            else:
                pass
    print("成功生产用户部门%d"%count)





if __name__=="__main__":
    print("执行函数名",sys.argv[1])
    if sys.argv[1] == "initdeparts":
        initdeparts()
    elif sys.argv[1] == "initsuperuser":
        initsuperuser()
    elif sys.argv[1] == "initposts":
        initposts()
    elif sys.argv[1] == "initusers":
        initusers()
    elif sys.argv[1] == "initrank13demands":
        initrank13demands()
    elif sys.argv[1] == "initrank13coe":
        initrank13coe()
    elif sys.argv[1] == "initcoe":
        initcoe()
    elif sys.argv[1] == "initcertificates":
        initcertificates()
    elif sys.argv[1] == "inituserspassword":
        inituserspassword()
    elif sys.argv[1] == "initusercertificates":
        initusercertificates()
    elif sys.argv[1] == "inituserdepart":
        inituserdepart()
    elif sys.argv[1] == "updateuserbaseinfo":
        updateuserbaseinfo()
    else:
        print(" 无效指令")


